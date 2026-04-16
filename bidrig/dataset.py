import json
import math
import statistics
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl


@dataclass(frozen=True)
class ProjectSample:
    bid_ann_guid: str
    projguid: str
    label: int
    project_info: Dict[str, Any]
    announcement: Dict[str, Any]
    bidders: List[Dict[str, Any]]
    indicators: Dict[str, Any]


def load_all_data_xlsx(excel_path: str) -> List[ProjectSample]:
    wb = openpyxl.load_workbook(excel_path)
    project_sheet = wb["项目信息"]
    ann_sheet = wb["招标公告"]
    bidder_sheet = wb["投标商数据"]

    project_rows = _read_sheet(project_sheet)
    ann_rows = _read_sheet(ann_sheet)
    bidder_rows = _read_sheet(bidder_sheet)

    project_by_guid = _dedupe_project_info_by_updatedate(project_rows, key="projguid")

    bidders_by_ann: Dict[str, List[Dict[str, Any]]] = {}
    for row in bidder_rows:
        bid_ann_guid = row.get("bid_ann_guid")
        if not bid_ann_guid:
            continue
        bidders_by_ann.setdefault(str(bid_ann_guid), []).append(row)

    samples: List[ProjectSample] = []
    for ann in ann_rows:
        bid_ann_guid = ann.get("bid_ann_guid")
        projguid = ann.get("projguid")
        if not bid_ann_guid or not projguid:
            continue

        bid_ann_guid = str(bid_ann_guid)
        projguid = str(projguid)

        bidders = bidders_by_ann.get(bid_ann_guid, [])
        if len(bidders) != 5:
            continue

        label = _to_int01(ann.get("is_collusion"))
        if label is None:
            label = _to_int01(bidders[0].get("is_collusion"))
        if label is None:
            label = _to_int01(project_by_guid.get(projguid, {}).get("is_collusion"))
        if label is None:
            label = 0

        selected_project_info = dict(project_by_guid.get(projguid, {}))
        selected_announcement = dict(ann)
        selected_bidders = [_normalize_bidder_row(b) for b in bidders]

        indicators = _compute_indicators(selected_bidders)

        samples.append(
            ProjectSample(
                bid_ann_guid=bid_ann_guid,
                projguid=projguid,
                label=int(label),
                project_info=selected_project_info,
                announcement=selected_announcement,
                bidders=selected_bidders,
                indicators=indicators,
            )
        )

    return samples


def _read_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, Any]]:
    headers = [c.value for c in sheet[1]]
    rows: List[Dict[str, Any]] = []
    for r in range(2, sheet.max_row + 1):
        row: Dict[str, Any] = {}
        for j, h in enumerate(headers, 1):
            if h is None:
                continue
            value = sheet.cell(row=r, column=j).value
            if isinstance(value, str):
                value = value.strip()
                if value and (value[0] in "[{" and value[-1] in "]}"):
                    parsed = _try_parse_json(value)
                    if parsed is not None:
                        value = parsed
            row[str(h)] = value
        rows.append(row)
    return rows


def _try_parse_json(text: str) -> Optional[Any]:
    try:
        return json.loads(text)
    except Exception:
        return None


def _dedupe_project_info_by_updatedate(rows: Iterable[Dict[str, Any]], key: str) -> Dict[str, Dict[str, Any]]:
    best: Dict[str, Tuple[Optional[datetime], Dict[str, Any]]] = {}
    for row in rows:
        guid = row.get(key)
        if not guid:
            continue
        guid = str(guid)
        updated_at = _parse_datetime(row.get("updatedate")) or _parse_datetime(row.get("creationdate"))
        if guid not in best:
            best[guid] = (updated_at, dict(row))
            continue
        prev_dt, _ = best[guid]
        if prev_dt is None and updated_at is not None:
            best[guid] = (updated_at, dict(row))
            continue
        if prev_dt is not None and updated_at is not None and updated_at > prev_dt:
            best[guid] = (updated_at, dict(row))
            continue
    return {k: v for k, (_, v) in best.items()}


def _parse_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    if isinstance(value, (int, float)):
        return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except Exception:
                continue
    return None


def _to_int01(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        if int(value) in (0, 1):
            return int(value)
        return 1 if value > 0 else 0
    if isinstance(value, str):
        text = value.strip()
        if text in ("0", "1"):
            return int(text)
        if text.lower() in ("yes", "true"):
            return 1
        if text.lower() in ("no", "false"):
            return 0
    return None


def _normalize_bidder_row(row: Dict[str, Any]) -> Dict[str, Any]:
    price = row.get("x_price")
    normalized_price = _to_float(price)
    wins = _to_int01(row.get("x_isqualified"))
    bid_time = row.get("versionnumber")

    normalized = dict(row)
    normalized["x_price"] = normalized_price
    normalized["x_isqualified"] = wins if wins is not None else row.get("x_isqualified")
    normalized["versionnumber"] = bid_time
    normalized["x_providername"] = row.get("x_providername")
    normalized["x_biddercontact"] = row.get("x_biddercontact")
    return normalized


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return None
        except Exception:
            pass
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace(",", "")
        if text.count(".") > 1:
            parts = text.split(".")
            text = parts[0] + "." + parts[1]
        try:
            return float(text)
        except Exception:
            return None
    return None


def _compute_indicators(bidders: List[Dict[str, Any]]) -> Dict[str, Any]:
    prices = [b.get("x_price") for b in bidders if isinstance(b.get("x_price"), (int, float))]
    wins = [b.get("x_isqualified") for b in bidders]

    ordered = sorted(
        [(b.get("x_price"), b) for b in bidders if isinstance(b.get("x_price"), (int, float))],
        key=lambda t: t[0],
    )
    ordered_bidders = [b for _, b in ordered]

    losing_prices = [
        b.get("x_price")
        for b in bidders
        if isinstance(b.get("x_price"), (int, float)) and b.get("x_isqualified") == 0
    ]

    rd = None
    cv_losing = None

    if len(ordered_bidders) >= 2:
        lowest = ordered_bidders[0].get("x_price")
        second_lowest = ordered_bidders[1].get("x_price")
        if isinstance(lowest, (int, float)) and isinstance(second_lowest, (int, float)):
            delta = second_lowest - lowest
            sigma = statistics.stdev(losing_prices) if len(losing_prices) >= 2 else 0.0
            rd = float("inf") if sigma == 0 else delta / sigma

    if losing_prices:
        mu = statistics.mean(losing_prices) if len(losing_prices) >= 1 else 0.0
        sigma = statistics.stdev(losing_prices) if len(losing_prices) >= 2 else 0.0
        cv_losing = float("inf") if mu == 0 else sigma / mu

    price_std = statistics.stdev(prices) if len(prices) >= 2 else 0.0
    price_mean = statistics.mean(prices) if prices else 0.0
    price_cv = float("inf") if price_mean == 0 else price_std / price_mean

    contacts = [str(b.get("x_biddercontact")) for b in bidders if b.get("x_biddercontact")]
    contact_dup_count = len(contacts) - len(set(contacts))

    phones: List[str] = []
    emails: List[str] = []
    for b in bidders:
        for p in b.get("registration_phone_numbers") or []:
            if p:
                phones.append(str(p))
        for e in b.get("registration_emails") or []:
            if e:
                emails.append(str(e))

    phone_dup_count = len(phones) - len(set(phones))
    email_dup_count = len(emails) - len(set(emails))

    return {
        "rd": rd,
        "cv_losing": cv_losing,
        "price_cv_all": price_cv,
        "contact_dup_count": contact_dup_count,
        "phone_dup_count": phone_dup_count,
        "email_dup_count": email_dup_count,
        "wins": wins,
    }

