import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import numpy as np
from sklearn.metrics import precision_recall_curve, auc

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 文件路径
method = 3
company = "zhipu"   # "zhipu" 或 "chat"
result_path = '.\\实验结果\\output_method{}_model-{}_masked.xlsx'.format(method, company)
label_path = '.\\数据\\testdata_masked.xlsx'

bad_bids = []
w1 = load_workbook(label_path)
ws1 = w1.active
bad_table = []
for cell in ws1['B']:
    if cell.value != None:
        bad_table.append(int(cell.value))
for i in range(len(bad_table)):
    if bad_table[i] == 1:
        bad_bids.append(i+1)  # 转换为0-based索引
print(f"异常招标索引: {bad_bids}, 共{len(bad_bids)}个")

# 检查文件是否存在
if not os.path.exists(result_path):
    print(f"文件 {result_path} 不存在")
    exit()

# 加载数据
workbook = load_workbook(result_path)
worksheet = workbook.active

# 获取数据行数
rows = worksheet.max_row

# 读取置信度分数（B列）
scores = []
for row in range(2, rows + 1):  # 假设第一行是表头，从第二行开始读取数据
    cell_value = worksheet[f'B{row}'].value
    scores.append(cell_value)

# 创建真实标签
# 1表示异常（bad_bids中指定的行），0表示正常
y_true = [0] * len(scores)
for idx in bad_bids:
    if 0 <= idx < len(y_true):  # 确保索引有效
        y_true[idx] = 1

# 计算PR曲线
precision, recall, thresholds = precision_recall_curve(y_true, scores)
for p, r, t in zip(precision, recall, thresholds):
    print(f"Precision: {p:.4f}, Recall: {r:.4f}, Threshold: {t:.4f}")
    pass

# 计算AUC-PR
pr_auc = auc(recall, precision)

# 计算F1分数
#f1_scores = 2 * (precision * recall) / (precision + recall)
f1_scores = np.array([])
for p, r in zip(precision, recall):
    if p + r == 0:
        f1_scores = np.append(f1_scores, 0)
    else:
        f1 = 2 * (p * r) / (p + r)
        f1_scores = np.append(f1_scores, f1)

# 找到最优F1分数及其对应的点
best_f1_idx = np.argmax(f1_scores)
best_f1 = f1_scores[best_f1_idx]
best_precision = precision[best_f1_idx]
best_recall = recall[best_f1_idx]

# 绘制PR曲线
plt.figure(figsize=(8, 6))
plt.plot(recall, precision, color='blue', lw=2, label=f'PR曲线 (AUC = {pr_auc:.2f})')

# 标注最优F1分数点
plt.scatter(best_recall, best_precision, color='red', s=20, label=f'最优F1点 (F1 = {best_f1:.2f})')
plt.annotate(f'F1 = {best_f1:.2f}\nRecall = {best_recall:.2f}\nPrecision = {best_precision:.2f}',
             xy=(best_recall, best_precision), xytext=(best_recall-0.1, best_precision + 0.1),
             arrowprops=dict(facecolor='black', shrink=0.05, width=1, headwidth=7),
             fontsize=10,
             bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5))

plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('召回率 (Recall)')
plt.ylabel('精确率 (Precision)')
plt.title(f'方法1(无提示)的PR曲线')
plt.legend(loc="best")
plt.grid(True)

# 显示图像
plt.show()