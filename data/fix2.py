from openpyxl import load_workbook
import json, random
import statistics
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import matplotlib.dates as mdates
import matplotlib.font_manager as fm

plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用黑体
plt.rcParams['axes.unicode_minus'] = False   # 解决负号显示问题

test_data = '.\\testdata.xlsx'

def get_RD(tb_group):

    # 分离出中标者和落败者的出价
    winning_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 1]
    losing_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 0]
    # 计算两个最低出价之差Δ
    if len(tb_group) >= 2:
        lowest_bid = tb_group[0]['x_price']
        second_lowest_bid = tb_group[1]['x_price']
        delta = second_lowest_bid - lowest_bid
    else:
        delta = 0  # 出价数量不足时无法计算

    # 计算落败竞价的标准差σ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0  # 只有一个落败出价时标准差为0
    else:
        sigma = 0

    # 计算比值Δ/σ
    if sigma != 0:
        ratio = delta / sigma
    else:
        ratio = float('inf')  # 如果标准差为0，比值设为无穷大
    return ratio

def get_CV(tb_group):
    # 分离出落败者的出价
    losing_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 0]

    # 计算落败竞价的标准差σ和算术平均数μ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
        mu = statistics.mean(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0
        mu = losing_bids[0]
    else:
        sigma = 0
        mu = 0

    # 计算变异系数CV=σ/μ
    if mu != 0:
        cv = sigma / mu
    else:
        cv = float('inf')  # 如果平均数为0，变异系数设为无穷大
    return cv

if __name__ == "__main__":
    # 加载原始Excel文件
    workbook = load_workbook(filename=test_data)
    worksheet = workbook.active

    xm_list, zb_list, tb_list = [], [], []
    connecter = ["projguid", "bid_ann_guid"]

    # 读取A列数据并处理
    for cell in worksheet['A']:
        if cell.value[0] != "{":
            continue
        dict_str = cell.value
        if dict_str:
            try:
                # 解析字典字符串
                data_dict = json.loads(dict_str)
                # 根据connecter分类
                if connecter[0] in data_dict and connecter[1] not in data_dict:
                    xm_list.append(data_dict)
                elif connecter[1] in data_dict and connecter[0] not in data_dict:
                    tb_list.append(data_dict)
                elif connecter[0] in data_dict and connecter[1] in data_dict:
                    zb_list.append(data_dict)
                else:
                    assert(False)
            except json.JSONDecodeError as e:
                print(f"第{cell.row}行解析失败: {e}")
                continue
    print(f"项目数量: {len(xm_list)}, 招标数量: {len(zb_list)}, 投标数量: {len(tb_list)}")

    evil_bids = []
    while len(evil_bids) < len(zb_list) // 2:
        bid = random.randint(0, len(zb_list) - 1)
        if bid not in evil_bids:
            evil_bids.append(bid)
    evil_bids.sort()
    print(f"异常招标索引: {evil_bids}, 数量: {len(evil_bids)}")

    new_tb_list = []

    for i in range(len(zb_list)):
        zb = zb_list[i]
        tb_group = []
        for tb in tb_list:
            if zb['bid_ann_guid'] == tb['bid_ann_guid']:
                tb_group.append(tb)
        tb_group.sort(key=lambda x: x['x_price'])

        # 确保每个招标有且只有一个中标者
        tb_group[0]['x_isqualified'] = 1
        for j in range(1, len(tb_group)):
            tb_group[j]['x_isqualified'] = 0

        ava_name_list = ["张家豪", "李思聪", "王子轩", "刘晨曦", "陈昊天", "杨宇轩", "赵龙", "黄梓涵", "周子涵", "吴俊杰",
                         "徐浩然", "孙宇航", "马俊杰", "朱子", "胡睿哲", "郭子豪", "何宇轩", "林浩然", "高子轩", "罗俊杰"]
        
        # 修改异常招标中投标商的数据
        if i in evil_bids:
            # 价格数值调整
            low_edge = random.randint(11,22) / 10
            print(f"处理第{i}个招标，目标RD区间: [{low_edge}, 3]")
            while True:
                rd = get_RD(tb_group)

                if rd >= low_edge and rd <= 3:
                    break
                elif rd < low_edge:
                    # 比值过小，增大Δ
                    if len(tb_group) >= 2:
                        tb_group[-1]['x_price'] -= 1
                elif rd > 3:
                    # 比值过大，减小Δ
                    if len(tb_group) >= 2:
                        tb_group[-1]['x_price'] += 1
                tb_group.sort(key=lambda x: x['x_price'])

            # 名字文本调整
            same_man = random.randint(1, len(tb_group) - 1)
            tb_group[same_man]['x_employee'] = tb_group[0]['x_employee']

            # 公司名文本调整
            pass

        else:
            # 价格调整
            while True:
                rd = get_RD(tb_group)
                if rd <= 2:
                    break
                else:
                    # 比值过大，减小Δ
                    if len(tb_group) >= 2:
                        tb_group[0]['x_price'] += 1000
                tb_group.sort(key=lambda x: x['x_price'])
            
            # 名字文本调整
            for tb_index in range(len(tb_group)):
                for tb_index_j in range(tb_index + 1, len(tb_group)):
                    tb_j = tb_group[tb_index_j]
                    while tb_j['x_employee'] == tb_group[tb_index]['x_employee']:
                        tb_group[tb_index_j]['x_employee'] = random.choice(ava_name_list)

            # 公司名文本调整
            ava_company_list = ["华夏科技智能建设有限公司", "三齐建设工程有限公司", "天河建设集团", "鹏程建筑工程有限公司", "新南国建设有限公司", "MAD建筑设计事务所"]
            company_list = []
            for tb in tb_group:
                comp_name = tb["x_providername"].split('//')
                for idx in range(len(comp_name)):
                    while comp_name[idx] in company_list:
                        comp_name[idx] = random.choice(ava_company_list)
                company_list.extend(comp_name)
                tb["x_providername"] = '//'.join(comp_name)

            # 电话文本调整
            phone_list = []
            for tb in tb_group:
                while tb['x_biddercontact'] in phone_list:
                    tb['x_biddercontact'] = '1' + ''.join([str(random.randint(0, 9)) for _ in range(10)])
                phone_list.append(tb['x_biddercontact'])

        new_tb_list.extend(tb_group)
    
    print(new_tb_list[:5])
    print(len(new_tb_list))

    new_xlsx = test_data.replace(".xlsx", "_mod2.xlsx")
    new_workbook = load_workbook(filename=test_data)
    new_worksheet = new_workbook.active
    flag = 0
    st = 0
    print(new_worksheet.max_row)
    for row in range(1, new_worksheet.max_row + 1):
        cell = new_worksheet.cell(row=row, column=1).value
        if cell[0] != "{":
            print(cell)
            st = row
            flag += 1
        elif flag == 3:
            new_worksheet.cell(row=row, column=1).value = json.dumps(new_tb_list.pop(0), ensure_ascii=False)
        elif flag == 2:
            new_worksheet.cell(row=row, column=2).value = 1 if row - st - 1 in evil_bids else 0
    new_workbook.save(filename=new_xlsx)