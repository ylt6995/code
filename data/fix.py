import json
from openpyxl import load_workbook, Workbook

# 加载原始Excel文件
workbook = load_workbook(filename='all_testdata.xlsx')
worksheet = workbook.active

tmp = []

for i in range(202, 697):
    cell = worksheet.cell(row=i, column=1)
    dict_str = cell.value
    data_dict = json.loads(dict_str)
    tmp.append(data_dict)
    if len(tmp) == 5:
        # 处理逻辑：当收集到5个字典时进行处理
        # 1. 统计x_isqualified为1的字典
        qualified_dicts = [d for d in tmp if d.get('x_isqualified') == 1]
        
        if len(qualified_dicts) == 1:
            # 如果只有一个字典的x_isqualified为1，保持不变
            pass
        elif len(qualified_dicts) > 1:
            # 如果有多个x_isqualified为1，选择x_price最小的那个保留为1，其余置为0
            # 找到x_price最小的字典
            min_price_dict = min(qualified_dicts, key=lambda x: x.get('x_price', float('inf')))
            
            # 将所有字典的x_isqualified置为0
            for d in tmp:
                d['x_isqualified'] = 0
            
            # 只将x_price最小的字典的x_isqualified置为1
            min_price_dict['x_isqualified'] = 1
        else:
            # 如果所有x_isqualified为0，选择x_price最小的那个置为1
            min_price_dict = min(tmp, key=lambda x: x.get('x_price', float('inf')))
            min_price_dict['x_isqualified'] = 1
        
        # 处理完成后，可以将结果写回Excel或进行其他操作
        # 这里只是示例，您可能需要根据实际需求调整
        for j, d in enumerate(tmp):
            # 将处理后的字典写回单元格
            worksheet.cell(row=i-4+j, column=1).value = json.dumps(d, ensure_ascii=False)
        
        # 清空tmp，准备下一批数据
        tmp = []

# 保存修改后的Excel文件
workbook.save(filename='all_testdata_modified.xlsx')