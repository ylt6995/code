from openpyxl import load_workbook
import json, random
import statistics
from faker import Faker
from datetime import date

test_data = '.\\all_data.xlsx'

fake = Faker('zh_CN')

def gen_fake_principals():
    fake_principals = [{} for _ in range(random.randint(2, 4))]
    position_list = ["General Manager", "Deputy General Manager", "Director", "Director"]
    for j in fake_principals:
        j["name"] = fake.name()
        j["position"] = position_list[fake_principals.index(j)]
        j["shareholding_ratio"] = random.randint(0, 10) / 100
    return fake_principals

def gen_fake_shareholders():
    fake_shareholders = [{} for _ in range(random.randint(2, 4))]
    for j in fake_shareholders:
        j["name"] = fake.name()
        j["shareholding_ratio"] = random.randint(0, 25) / 100
        j["subscribed_capital"] = random.randint(0, 100) * 10000
    return fake_shareholders

def gen_fake_registration_time():
    start_date = date(2000, 1, 1)
    end_date = date(2020, 12, 31)
    random_date = fake.date_between(start_date=start_date, end_date=end_date)
    return f"{random_date.year}-{random_date.month:02d}-{random_date.day:02d}"

def gen_fake_registration_time_narrower():
    start_date = date(2015, 1, 1)
    end_date = date(2017, 12, 31)
    random_date = fake.date_between(start_date=start_date, end_date=end_date)
    return f"{random_date.year}-{random_date.month:02d}-{random_date.day:02d}"

def gen_fake_phone_numbers():
    ph_list = []
    for _ in range(random.randint(1, 3)):
        ph_list.append(fake.phone_number())
    return ph_list

def gen_fake_emails():
    em_list = []
    for _ in range(random.randint(1, 3)):
        em_list.append(fake.email())
    return em_list


if __name__ == "__main__":
    # 加载原始Excel文件
    workbook = load_workbook(filename=test_data)
    worksheet = workbook['投标商数据']

    tb_list = []
    first_row = list(worksheet['1'])
    rows = list(worksheet['A'])
    for i in range(1, len(rows)):
        row_data = {}
        for j in range(len(first_row)):
            row_data[first_row[j].value] = list(worksheet[str(i+1)])[j].value
            if isinstance(row_data[first_row[j].value], str):
              
                if "[" in row_data[first_row[j].value] or "{" in row_data[first_row[j].value]:
                    row_data[first_row[j].value] = eval(row_data[first_row[j].value])
        tb_list.append(row_data)
    
    bad_table = []
    for i in range(0, len(tb_list), 5):
        bad_table.append(tb_list[i]['is_collusion'])
    print(f"异常招标索引: {[i+1 for i, val in enumerate(bad_table) if val == 1]}, 共{sum(bad_table)}个")

    labels = []
    for i in first_row:
        labels.append(i.value)
    if labels[-1] != "is_collusion":
        print("异常：最后一列不是is_collusion")
        exit("0")
    if "principals" not in labels:
        worksheet.cell(row=1, column=len(labels)+1, value="principals")
    if "shareholders" not in labels:
        worksheet.cell(row=1, column=len(labels)+2, value="shareholders")
    if "registration_time" not in labels:
        worksheet.cell(row=1, column=len(labels)+3, value="registration_time")
    if "registration_place" not in labels:
        worksheet.cell(row=1, column=len(labels)+4, value="registration_place")
    if "registration_phone_numbers" not in labels:
        worksheet.cell(row=1, column=len(labels)+5, value="registration_phone_numbers")
    if "registration_emails" not in labels:
        worksheet.cell(row=1, column=len(labels)+6, value="registration_emails")

    count = 0
    for i in range(0, len(tb_list), 5):
        if i > count:
            count += len(tb_list) // 10
            print(f"已处理{i}条数据")
        if tb_list[i]["is_collusion"] == 1:
            group_a = gen_fake_principals()
            address_a = fake.address()
            phone_numbers_a = gen_fake_phone_numbers()
            email_a = gen_fake_emails()
            worksheet.cell(row=i+2, column=len(labels)+1, value=json.dumps(group_a, ensure_ascii=False))
            worksheet.cell(row=i+2, column=len(labels)+2, value=json.dumps(gen_fake_shareholders(), ensure_ascii=False))
            worksheet.cell(row=i+2, column=len(labels)+3, value=gen_fake_registration_time_narrower())
            worksheet.cell(row=i+2, column=len(labels)+4, value=address_a)
            worksheet.cell(row=i+2, column=len(labels)+5, value=json.dumps(phone_numbers_a, ensure_ascii=False))
            worksheet.cell(row=i+2, column=len(labels)+6, value=json.dumps(email_a, ensure_ascii=False))

            group_b = gen_fake_shareholders()
            group_b[1]["name"] = group_a[0]["name"]
            worksheet.cell(row=i+3, column=len(labels)+1, value=json.dumps(gen_fake_principals(), ensure_ascii=False))
            worksheet.cell(row=i+3, column=len(labels)+2, value=json.dumps(group_b, ensure_ascii=False))
            worksheet.cell(row=i+3, column=len(labels)+3, value=gen_fake_registration_time_narrower())
            worksheet.cell(row=i+3, column=len(labels)+4, value=fake.address())
            worksheet.cell(row=i+3, column=len(labels)+5, value=json.dumps(gen_fake_phone_numbers(), ensure_ascii=False))
            worksheet.cell(row=i+3, column=len(labels)+6, value=json.dumps(gen_fake_emails(), ensure_ascii=False))

            group_c = gen_fake_shareholders()
            group_c[1]["name"] = group_a[1]["name"]
            worksheet.cell(row=i+4, column=len(labels)+1, value=json.dumps(gen_fake_principals(), ensure_ascii=False))
            worksheet.cell(row=i+4, column=len(labels)+2, value=json.dumps(group_c, ensure_ascii=False))
            worksheet.cell(row=i+4, column=len(labels)+3, value=gen_fake_registration_time_narrower())
            worksheet.cell(row=i+4, column=len(labels)+4, value=address_a)
            worksheet.cell(row=i+4, column=len(labels)+5, value=json.dumps(gen_fake_phone_numbers(), ensure_ascii=False))
            worksheet.cell(row=i+4, column=len(labels)+6, value=json.dumps(gen_fake_emails(), ensure_ascii=False))

            group_d = gen_fake_principals()
            group_d[0]["name"] = group_b[0]["name"]
            worksheet.cell(row=i+5, column=len(labels)+1, value=json.dumps(group_d, ensure_ascii=False))
            worksheet.cell(row=i+5, column=len(labels)+2, value=json.dumps(gen_fake_shareholders(), ensure_ascii=False))
            worksheet.cell(row=i+5, column=len(labels)+3, value=gen_fake_registration_time_narrower())
            worksheet.cell(row=i+5, column=len(labels)+4, value=fake.address())
            worksheet.cell(row=i+5, column=len(labels)+5, value=json.dumps(phone_numbers_a[:1], ensure_ascii=False))
            worksheet.cell(row=i+5, column=len(labels)+6, value=json.dumps(gen_fake_emails(), ensure_ascii=False))

            worksheet.cell(row=i+6, column=len(labels)+1, value=json.dumps(group_a[:2], ensure_ascii=False))
            worksheet.cell(row=i+6, column=len(labels)+2, value=json.dumps(group_b[:1], ensure_ascii=False))
            worksheet.cell(row=i+6, column=len(labels)+3, value=gen_fake_registration_time_narrower())
            worksheet.cell(row=i+6, column=len(labels)+4, value=fake.address())
            worksheet.cell(row=i+6, column=len(labels)+5, value=json.dumps(gen_fake_phone_numbers(), ensure_ascii=False))
            worksheet.cell(row=i+6, column=len(labels)+6, value=json.dumps(email_a[:1], ensure_ascii=False))

        elif tb_list[i]["is_collusion"] == 0:
            for j in range(2, 7):
                worksheet.cell(row=i+j, column=len(labels)+1, value=json.dumps(gen_fake_principals(), ensure_ascii=False))
                worksheet.cell(row=i+j, column=len(labels)+2, value=json.dumps(gen_fake_shareholders(), ensure_ascii=False))
                worksheet.cell(row=i+j, column=len(labels)+3, value=gen_fake_registration_time())
                worksheet.cell(row=i+j, column=len(labels)+4, value=fake.address())
                worksheet.cell(row=i+j, column=len(labels)+5, value=json.dumps(gen_fake_phone_numbers(), ensure_ascii=False))
                worksheet.cell(row=i+j, column=len(labels)+6, value=json.dumps(gen_fake_emails(), ensure_ascii=False))

        else:
            raise ValueError(f"第{i+1}条数据异常，is_collusion值为{tb_list[i]['is_collusion']}")

    workbook.save(filename=test_data)