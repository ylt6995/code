from openpyxl import load_workbook
import json, random
import statistics

test_data = '.\\testdata.xlsx'

def get_RD(tb_group):

    # 分离出中标者和落败者的出价
    winning_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 1]
    losing_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 0]
    # 计算两个最低出价之差Δ
    if len(tb_group) >= 2:
        lowest_bid = tb_group[0]['x_price']
        second_lowest_bid = tb_group[1]['x_price']
        delta = second_lowest_bid - lowest_bid
    else:
        delta = 0  # 出价数量不足时无法计算

    # 计算落败竞价的标准差σ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0  # 只有一个落败出价时标准差为0
    else:
        sigma = 0

    # 计算比值Δ/σ
    if sigma != 0:
        ratio = delta / sigma
    else:
        ratio = float('inf')  # 如果标准差为0，比值设为无穷大
    return ratio

def get_CV(tb_group):
    # 分离出落败者的出价
    losing_bids = [bid['x_price'] for bid in tb_group if bid['x_isqualified'] == 0]

    # 计算落败竞价的标准差σ和算术平均数μ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
        mu = statistics.mean(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0
        mu = losing_bids[0]
    else:
        sigma = 0
        mu = 0

    # 计算变异系数CV=σ/μ
    if mu != 0:
        cv = sigma / mu
    else:
        cv = float('inf')  # 如果平均数为0，变异系数设为无穷大
    return cv

if __name__ == "__main__":
    # 加载原始Excel文件
    workbook = load_workbook(filename=test_data)
    worksheet = workbook.active

    xm_list, zb_list, tb_list = [], [], []
    connecter = ["projguid", "bid_ann_guid"]

    # 读取A列数据并处理
    for cell in worksheet['A']:
        if cell.value[0] != "{":
            continue
        dict_str = cell.value
        if dict_str:
            try:
                # 解析字典字符串
                data_dict = json.loads(dict_str)
                # 根据connecter分类
                if connecter[0] in data_dict and connecter[1] not in data_dict:
                    xm_list.append(data_dict)
                elif connecter[1] in data_dict and connecter[0] not in data_dict:
                    tb_list.append(data_dict)
                elif connecter[0] in data_dict and connecter[1] in data_dict:
                    zb_list.append(data_dict)
                else:
                    assert(False)
            except json.JSONDecodeError as e:
                print(f"第{cell.row}行解析失败: {e}")
                continue
    print(f"项目数量: {len(xm_list)}, 招标数量: {len(zb_list)}, 投标数量: {len(tb_list)}")

    evil_bids = []
    while len(evil_bids) < len(zb_list) // 2:
        bid = random.randint(0, len(zb_list) - 1)
        if bid not in evil_bids:
            evil_bids.append(bid)
    evil_bids.sort()
    print(f"异常招标索引: {evil_bids}, 数量: {len(evil_bids)}")

    new_tb_list = []

    for i in range(len(zb_list)):
        zb = zb_list[i]
        tb_group = []
        for tb in tb_list:
            if zb['bid_ann_guid'] == tb['bid_ann_guid']:
                tb_group.append(tb)
        tb_group.sort(key=lambda x: x['x_price'])

        # 确保每个招标有且只有一个中标者
        tb_group[0]['x_isqualified'] = 1
        for j in range(1, len(tb_group)):
            tb_group[j]['x_isqualified'] = 0

        # 固定员工名称列表
        fixed_names = ["张三", "李四", "王五", "赵六", "朱八"]
        # 固定公司名称列表
        fixed_companies = ["公司A", "公司B", "公司C", "公司D", "公司E"]
        
        # 修改异常招标中投标商的数据
        if i in evil_bids:
            # 价格数值调整
            low_edge = random.randint(11,22) / 10
            print(f"处理第{i}个招标，目标RD区间: [{low_edge}, 3]")
            while True:
                rd = get_RD(tb_group)

                if rd >= low_edge and rd <= 3:
                    break
                elif rd < low_edge:
                    # 比值过小，增大Δ
                    if len(tb_group) >= 2:
                        tb_group[-1]['x_price'] -= 1
                elif rd > 3:
                    # 比值过大，减小Δ
                    if len(tb_group) >= 2:
                        tb_group[-1]['x_price'] += 1
                tb_group.sort(key=lambda x: x['x_price'])

        else:
            # 价格调整
            while True:
                rd = get_RD(tb_group)
                if rd <= 2:
                    break
                else:
                    # 比值过大，减小Δ
                    if len(tb_group) >= 2:
                        tb_group[0]['x_price'] += 1000
                tb_group.sort(key=lambda x: x['x_price'])
        
        # 统一处理：固定员工名称、公司名称和电话号码
        for idx, tb in enumerate(tb_group):
            # 固定员工名称
            if idx < len(fixed_names):
                tb['x_employee'] = fixed_names[idx]
                tb['x_bidder'] = fixed_names[idx]
            else:
                tb['x_employee'] = f"员工{idx+1}"
                tb['x_bidder'] = f"员工{idx+1}"
            
            # 固定公司名称
            if idx < len(fixed_companies):
                tb["x_providername"] = fixed_companies[idx]
            else:
                tb["x_providername"] = f"公司{chr(65+idx)}"
            
            # 固定电话号码格式
            tb['x_biddercontact'] = '1' + ''.join(['X'] * 10)

        new_tb_list.extend(tb_group)
    
    print(new_tb_list[:5])     # 输出前5条数据进行检查效果
    print(f"总投标数量: {len(new_tb_list)}")

    new_xlsx = test_data.replace(".xlsx", "_masked.xlsx")
    new_workbook = load_workbook(filename=test_data)
    new_worksheet = new_workbook.active
    flag = 0
    st = 0
    print(f"原始数据总行数: {new_worksheet.max_row}")
    for row in range(1, new_worksheet.max_row + 1):
        cell = new_worksheet.cell(row=row, column=1).value
        if cell[0] != "{":
            print(f"第{row}行数据异常: {cell}")
            st = row
            flag += 1
        elif flag == 3:
            new_worksheet.cell(row=row, column=1).value = json.dumps(new_tb_list.pop(0), ensure_ascii=False)
        elif flag == 2:
            new_worksheet.cell(row=row, column=2).value = 1 if row - st - 1 in evil_bids else 0
    new_workbook.save(filename=new_xlsx)