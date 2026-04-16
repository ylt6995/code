import openpyxl
import json
import openai
from time import sleep, time
import os
from datetime import datetime
import re
import statistics

company = os.getenv("LLM_PROVIDER", "zhipu")
model_list = {"zhipu": os.getenv("ZHIPU_MODEL", "glm-4.7-flash"), "chat": os.getenv("GPT_MODEL", "gpt-4o-mini")}

zhipu_api_key = os.getenv("ZHIPU_API_KEY", "")
zhipu_base_url = os.getenv("ZHIPU_BASE_URL", "https://open.bigmodel.cn/api/paas/v4")
chat_api_key = os.getenv("GPT_API_KEY", "") or os.getenv("OPENAI_API_KEY", "")
chat_base_url = os.getenv("GPT_BASE_URL", "") or os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")

ak = eval(f"{company}_api_key")
bu = eval(f"{company}_base_url")
model_name = model_list[company]

method = 4 # 1, 2, 3分别对应三种提示词设计
delay = 60  # 每次请求的延迟时间，单位秒

# Excel文件路径
excel_path = '.\\data\\all_data.xlsx'
result_path = '.\\实验结果\\output_method{}_model-{}_0314.xlsx'.format(method, company)

# 加载Excel文件
wb = openpyxl.load_workbook(excel_path)
sheet3 = wb['投标商数据']
sheet2 = wb['招标公告']

xm_list, zb_list, tb_list = [], [], []

def read_sheet(sheet):
    first_row = list(sheet['1'])
    content = []
    rows = list(sheet['A'])
    for i in range(1, len(rows)):
        row_data = {}
        for j in range(len(first_row)):
            row_data[first_row[j].value] = list(sheet[str(i+1)])[j].value
            if isinstance(row_data[first_row[j].value], str):
                if "[" in row_data[first_row[j].value] or "{" in row_data[first_row[j].value]:
                    try:
                        row_data[first_row[j].value] = json.loads(row_data[first_row[j].value])
                    except:
                        raise Exception(f"第{i+1}行第{j+1}列数据解析失败，内容: {row_data[first_row[j].value]}")
        content.append(row_data)
    return content

zb_list = read_sheet(sheet2)
tb_list = read_sheet(sheet3)

bad_table = []
for i in zb_list:
    bad_table.append(i['is_collusion'])
print(f"异常招标索引: {[i+1 for i, val in enumerate(bad_table) if val == 1]}, 共{sum(bad_table)}个")

def get_RD(tb_group):

    # 分离出中标者和落败者的出价
    winning_bids = [bid['price'] for bid in tb_group if bid['winsTheBid'] == 1]
    losing_bids = [bid['price'] for bid in tb_group if bid['winsTheBid'] == 0]
    # 计算两个最低出价之差Δ
    if len(tb_group) >= 2:
        lowest_bid = tb_group[0]['price']
        second_lowest_bid = tb_group[1]['price']
        delta = second_lowest_bid - lowest_bid
    else:
        delta = 0  # 出价数量不足时无法计算

    # 计算落败竞价的标准差σ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0  # 只有一个落败出价时标准差为0
    else:
        sigma = 0

    # 计算比值Δ/σ
    if sigma != 0:
        ratio = delta / sigma
    else:
        ratio = float('inf')  # 如果标准差为0，比值设为无穷大
    return ratio

def get_CV(tb_group):
    # 分离出落败者的出价
    losing_bids = [bid['price'] for bid in tb_group if bid['winsTheBid'] == 0]

    # 计算落败竞价的标准差σ和算术平均数μ
    if len(losing_bids) >= 2:
        sigma = statistics.stdev(losing_bids)
        mu = statistics.mean(losing_bids)
    elif len(losing_bids) == 1:
        sigma = 0
        mu = losing_bids[0]
    else:
        sigma = 0
        mu = 0

    # 计算变异系数CV=σ/μ
    if mu != 0:
        cv = sigma / mu
    else:
        cv = float('inf')  # 如果平均数为0，变异系数设为无穷大
    return cv

def check_collusion(prompt):
    sleep(delay)  # 避免请求过于频繁
    client = openai.OpenAI(api_key = ak, base_url = bu)
    while True:
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[{"role": "system", "content": "You are an expert in procurement auditing and bid rigging detection."},
                          {"role": "user", "content": prompt}],
                timeout=300.0,  # 5分钟超时
                temperature=0.1,
                top_p=0.9,
                max_tokens=10000
            )
            return response.choices[0].message.content
        except Exception as e:
            # 检查是否是超时异常
            if "timeout" in str(e).lower() or "timed out" in str(e).lower():
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 请求超时，正在重试...")
                sleep(delay)  # 再次延迟，避免频繁请求
                continue
            else:
                # 其他异常，直接抛出
                raise

# 检查结果文件是否存在
processed_bid_ann_guids = set()
pipeline3 = []
pipeline2 = []

if os.path.exists(result_path):
    # 读取已处理的bid_ann_guid
    result_wb = openpyxl.load_workbook(result_path)
    result_ws = result_wb.active
    for row in result_ws.iter_rows(min_row=2, max_row=result_ws.max_row, min_col=1, max_col=1):
        bid_ann_guid = row[0].value
        if bid_ann_guid:
            processed_bid_ann_guids.add(bid_ann_guid)
            if result_ws.cell(row=row[0].row, column=2).value != None:
                pipeline3.append(bid_ann_guid)
            if result_ws.cell(row=row[0].row, column=6).value != None:
                pipeline2.append(bid_ann_guid)

else:
    # 创建新的结果文件
    result_wb = openpyxl.Workbook()
    result_ws = result_wb.active
    # 设置表头
    headers = ["bid_ann_guid", "collusionSuspicionScore", "riskLevel", "keyEvidence"]
    for col, header in enumerate(headers, 1):
        result_ws.cell(row=1, column=col, value=header)

# 计算总项目数和已完成项目数
total_projects = len(zb_list)
completed_projects = len(processed_bid_ann_guids)

# 打印提示词，仅一次
prompt_show_flag = True
prompt_show_flag2 = True

# 遍历每个招标，提取对应的五个投标
for zb in zb_list:
    current_start_time = time()
    bid_ann_guid = zb['bid_ann_guid']
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 检测结果:\n")
    # 跳过已处理的项目
    #if bid_ann_guid in processed_bid_ann_guids:
    #    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 已处理，跳过 | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)")
    #    continue
    
    # tb_group = [tb for tb in tb_list if tb['bid_ann_guid'] == bid_ann_guid]
    tb_group = []
    tb_group2 = []
    for tb in tb_list:
        if tb['bid_ann_guid'] == bid_ann_guid:
            # 选择特定的字段
            selected_tb = {}
            selected_tb['providername'] = f"company_{len(tb_group)+1}"  # 使用占位符公司名称
            selected_tb['price'] = tb.get('x_price', -1)
            if type(selected_tb['price']) == str:
                if selected_tb['price'].count(".") > 1:
                    tmp=selected_tb['price'].split(".")
                    selected_tb['price'] = float(tmp[0] + "." + tmp[1])
            selected_tb['bid_time'] = tb.get('versionnumber', -1)
            selected_tb['winsTheBid'] = tb.get('x_isqualified', -1)
            selected_tb2 = {}
            selected_tb2['companyname'] = f"company_{len(tb_group)+1}"  # 使用占位符公司名称
            selected_tb2['principals'] = tb.get('principals', [])
            selected_tb2['shareholders'] = tb.get('shareholders', [])
            selected_tb2['registration_time'] = tb.get('registration_time', "")
            selected_tb2['registration_place'] = tb.get('registration_place', "")
            selected_tb2['registration_phone_numbers'] = tb.get('registration_phone_numbers', [])
            selected_tb2['registration_emails'] = tb.get('registration_emails', [])
            tb_group.append(selected_tb)
            tb_group2.append(selected_tb2)
    if len(tb_group) != 5 or len(tb_group2) != 5:
        assert(False)
    
    if bid_ann_guid not in pipeline3:

        # 计算指标
        RD = get_RD(tb_group)
        CV = get_CV(tb_group)
        indicators = f"RD={RD:.2f}, CV={CV:.2f}"

        # 构造提示词
        prompt = "\n"
        for idx, tb in enumerate(tb_group, 1):
            prompt += f"Bidding information{idx}: {json.dumps(tb, ensure_ascii=False)}\n"

        # 方法4
        prompt4 = f"""Core Task:
        You need to analyze the provided bidder data for a single tender project and perform a comprehensive collusion risk assessment.

    Original Data: ```{prompt}```

    Important Indicators: {indicators}

    Analysis Notes:
        1. Is there an abnormally low bid alongside a cluster of closely priced high bids? Do the bid prices show a mathematical relationship or consistent gaps?
        2. Are there patterns directly indicating collusion and bid-rigging cartels?
        3. You can use the RD and CV indicators to help you analyze the bidder data. RD is the ratio of the difference between the lowest and second-lowest bid to the standard deviation of losing bids. CV is the coefficient of variation of losing bids. 
        4. Generally, a higher RD and a lower CV may indicate a higher risk of collusion. An RD of approximately 1 indicates that there is no difference in the bidding behavior of the winner and the rest of the bidders. An RD (much) larger than 1 indicates, however, that cover bidding may have taken place.

    Output Format Requirements:
        Provide your analysis in the following structured JSON format:
    ```
    {{
        "is_the_bidding_prices_abnormally_concentrated": ["Yes" or "No"],
        "is_the_winning_quotation_significantly_lower_than_other_quotations": ["Yes" or "No"],
        "is_there_any_direct_evidence_of_collusion": ["Yes" or "No"]
    }}
    ```
        """
        
        prompt = eval(f"prompt{method}")

        # 打印提示词，仅一次
        if prompt_show_flag:
            print(f"提示词示例 (方法{method}):\n{prompt}\n")
            prompt_show_flag = False


        # 检测是否存在围标串标
        result = check_collusion(prompt)
        print(f"{result}")
        # 解析结果
        try:
            # 提取JSON部分
            while True:
                try:
                    json_match = re.search(r'\{.*\}', result, re.DOTALL)
                    if json_match:
                        result_json = json.loads(json_match.group())
                    else:
                        # 如果没有找到JSON，尝试直接解析
                        result_json = json.loads(result)
                except:
                    pass
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析JSON失败，等待120秒后重试")
                    sleep(120)
                    result = check_collusion(prompt)
                else:
                    break
            
            # 提取数据
            is_abnormally_concentrated = result_json.get("is_the_bidding_prices_abnormally_concentrated", "-1")
            if "Yes" in is_abnormally_concentrated:
                is_abnormally_concentrated = 1
            elif "No" in is_abnormally_concentrated:
                is_abnormally_concentrated = 0 
            else:
                is_abnormally_concentrated = -1
            is_winning_quotation_lower = result_json.get("is_the_winning_quotation_significantly_lower_than_other_quotations", "-1")
            if "Yes" in is_winning_quotation_lower:
                is_winning_quotation_lower = 1
            elif "No" in is_winning_quotation_lower:
                is_winning_quotation_lower = 0 
            else:
                is_winning_quotation_lower = -1
            is_collusion = result_json.get("is_there_any_direct_evidence_of_collusion", "-1")
            if "Yes" in is_collusion:   
                is_collusion = 1
            elif "No" in is_collusion:
                is_collusion = 0 
            else:
                is_collusion = -1

            # 计算处理耗时（从上一行写入到这一行写入之间的时间差）
            current_end_time = time()
            processing_time = current_end_time - current_start_time - delay  # 减去请求延迟时间
            
            # 写入结果到Excel
            next_row = result_ws.max_row + 1
            result_ws.cell(row=next_row, column=1, value=bid_ann_guid)
            result_ws.cell(row=next_row, column=2, value=is_abnormally_concentrated)
            result_ws.cell(row=next_row, column=3, value=is_winning_quotation_lower)
            result_ws.cell(row=next_row, column=4, value=is_collusion)

            result_ws.cell(row=next_row, column=8, value=str(prompt))
            result_ws.cell(row=next_row, column=9, value=bad_table.pop(0))
            
            # 保存结果文件
            result_wb.save(result_path)
            completed_projects += 1
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 结果已保存到{result_path} | 处理耗时: {processing_time:.2f}秒(不计{delay}秒延迟) | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)\n")
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析结果失败: {e} | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)")
            #print(json_match.group())
            assert(False)

    if bid_ann_guid not in pipeline2:
        p2 = "\n"
        for idx, tb in enumerate(tb_group2, 1):
            p2 += f"Bidding company information{idx}: {json.dumps(tb, ensure_ascii=False)}\n"
        pp2 = f"""Core Task:
        You need to analyze the provided bidder data for a single tender project and perform a comprehensive collusion risk assessment.

    Original Data: ```{p2}```

    Analysis Notes:
        1. Is there any cross appointment between important personnel of different companies?
        2. Are there similar situations in the time and place of registration of different companies?
        3. Do phone numbers and emails from different companies duplicate?
        
    Output Format Requirements:
        Provide your analysis in the following structured JSON format:
    ```
    {{
        "is_cross_appointment": ["Yes" or "No"],
        "is_registration_time_and_place_similar": ["Yes" or "No"],
        "is_phone_number_or_email_duplicate": ["Yes" or "No"]
    }}
    ```
    """
        if prompt_show_flag2:
            print(f"提示词示例 (pipe2):\n{pp2}\n")
            prompt_show_flag2 = False

        result2 = check_collusion(pp2)
        print(f"result2: {result2}\n")
        # 解析结果
        try:
            # 提取JSON部分
            while True:
                try:
                    json_match = re.search(r'\{.*\}', result2, re.DOTALL)
                    if json_match:
                        result_json = json.loads(json_match.group())
                    else:
                        # 如果没有找到JSON，尝试直接解析
                        result_json = json.loads(result2)
                except:
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析JSON失败，等待120秒后重试")
                    sleep(120)
                    result = check_collusion(pp2)
                else:
                    break
            
            # 提取数据
            is_cross_appointment = result_json.get("is_cross_appointment", "-1")
            if "Yes" in is_cross_appointment:
                is_cross_appointment = 1
            elif "No" in is_cross_appointment:
                is_cross_appointment = 0 
            else:
                is_cross_appointment = -1
            is_registration_time_and_place_similar = result_json.get("is_registration_time_and_place_similar", "-1")
            if "Yes" in is_registration_time_and_place_similar:
                is_registration_time_and_place_similar = 1
            elif "No" in is_registration_time_and_place_similar:
                is_registration_time_and_place_similar = 0 
            else:
                is_registration_time_and_place_similar = -1
            is_phone_number_or_email_duplicate = result_json.get("is_phone_number_or_email_duplicate", "-1")
            if "Yes" in is_phone_number_or_email_duplicate:   
                is_phone_number_or_email_duplicate = 1
            elif "No" in is_phone_number_or_email_duplicate:
                is_phone_number_or_email_duplicate = 0 
            else:
                is_phone_number_or_email_duplicate = -1

            # 计算处理耗时（从上一行写入到这一行写入之间的时间差）
            current_end_time = time()
            processing_time = current_end_time - current_start_time - delay  # 减去请求延迟时间
            
            # 写入结果到Excel
            next_row = pipeline3.index(bid_ann_guid) + 2

            result_ws.cell(row=next_row, column=5, value=is_cross_appointment)
            result_ws.cell(row=next_row, column=6, value=is_registration_time_and_place_similar)
            result_ws.cell(row=next_row, column=7, value=is_phone_number_or_email_duplicate)
            
            # 保存结果文件
            result_wb.save(result_path)
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 结果已保存到{result_path} | 处理耗时: {processing_time:.2f}秒(不计{delay}秒延迟) | 进度: {next_row-1}/{total_projects} ({(next_row-1)/total_projects*100:.1f}%)\n")
            
        except Exception as e:
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析结果失败: {e} | 进度: {next_row-1}/{total_projects} ({(next_row-1)/total_projects*100:.1f}%)")
            #print(json_match.group())
            assert(False)

# 关闭结果文件
result_wb.close()
