import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import numpy as np
from sklearn.metrics import precision_recall_curve, auc

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 文件路径
method = 1
result_path = '.\\实验结果\\output_method{}_gpt.xlsx'.format(method)
result_path2 = '.\\实验结果\\output_method{}.xlsx'.format(method)

bad_bids = [5, 7, 20, 34, 46, 54, 65, 76, 88, 99]
for i in range(len(bad_bids)):
    bad_bids[i] -= 1  # 调整为0-based索引

# 检查文件是否存在
for path in [result_path, result_path2]:
    if not os.path.exists(path):
        print(f"文件 {path} 不存在")
        exit()

# 定义函数来加载数据并计算PR曲线
def process_data(file_path):
    # 加载数据
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    # 获取数据行数
    rows = worksheet.max_row
    
    # 读取置信度分数（B列）
    scores = []
    for row in range(2, rows + 1):  # 假设第一行是表头，从第二行开始读取数据
        cell_value = worksheet[f'B{row}'].value
        scores.append(cell_value)
    
    # 创建真实标签
    # 1表示异常（bad_bids中指定的行），0表示正常
    y_true = [0] * len(scores)
    for idx in bad_bids:
        if 0 <= idx < len(y_true):  # 确保索引有效
            y_true[idx] = 1
    
    # 计算PR曲线
    precision, recall, thresholds = precision_recall_curve(y_true, scores)
    
    # 计算AUC-PR
    pr_auc = auc(recall, precision)
    
    # 计算F1分数
    f1_scores = np.array([])
    for p, r in zip(precision, recall):
        if p + r == 0:
            f1_scores = np.append(f1_scores, 0)
        else:
            f1 = 2 * (p * r) / (p + r)
            f1_scores = np.append(f1_scores, f1)
    
    # 找到最优F1分数及其对应的点
    best_f1_idx = np.argmax(f1_scores)
    best_f1 = f1_scores[best_f1_idx]
    best_precision = precision[best_f1_idx]
    best_recall = recall[best_f1_idx]
    
    return precision, recall, pr_auc, best_precision, best_recall, best_f1

# 处理第一组数据
precision1, recall1, pr_auc1, best_precision1, best_recall1, best_f11 = process_data(result_path)

# 处理第二组数据
precision2, recall2, pr_auc2, best_precision2, best_recall2, best_f12 = process_data(result_path2)

# 绘制PR曲线
plt.figure(figsize=(8, 6))

# 绘制第一组PR曲线
plt.plot(recall1, precision1, color='blue', lw=2, label=f'GPT, AUC = {pr_auc1:.2f}')
plt.scatter(best_recall1, best_precision1, color='blue', s=30, marker='o')
plt.annotate(f'F1 = {best_f11:.2f}',
             xy=(best_recall1, best_precision1), xytext=(best_recall1+0.05, best_precision1+0.05),
             arrowprops=dict(facecolor='blue', shrink=0.05, width=1, headwidth=7),
             fontsize=10,
             bbox=dict(boxstyle='round,pad=0.5', fc='lightblue', alpha=0.5))

# 绘制第二组PR曲线
plt.plot(recall2, precision2, color='red', lw=2, label=f'GLM, AUC = {pr_auc2:.2f}')
plt.scatter(best_recall2, best_precision2, color='red', s=30, marker='o')
plt.annotate(f'F1 = {best_f12:.2f}',
             xy=(best_recall2, best_precision2), xytext=(best_recall2-0.05, best_precision2+0.1),
             arrowprops=dict(facecolor='red', shrink=0.05, width=1, headwidth=7),
             fontsize=10,
             bbox=dict(boxstyle='round,pad=0.5', fc='lightcoral', alpha=0.5))

plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('召回率 (Recall)')
plt.ylabel('精确率 (Precision)')
plt.title(f'GPT和GLM两组实验结果的PR曲线对比')
plt.legend(loc="best")
plt.grid(True)

# 显示图像
plt.show()