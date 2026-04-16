from openpyxl import load_workbook
import matplotlib.pyplot as plt

import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    confusion_matrix, accuracy_score, f1_score,
    roc_curve, roc_auc_score
)

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

if __name__ == "__main__":
    # method=1: 使用BCDEFG列
    # method=2: 只使用BCD列  
    # method=3: 只使用EFG列
    method = 3 
    
    # 加载原始Excel文件
    workbook = load_workbook(filename='.\实验结果\output_method4_model-zhipu_0314.xlsx')
    worksheet = workbook.active
    
    # 读取数据
    X = [] 
    y = [] 
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        b_val = row[1]  # B列
        c_val = row[2]  # C列
        d_val = row[3]  # D列
        e_val = row[4]  # E列
        f_val = row[5]  # F列
        g_val = row[6]  # G列
        i_val = row[8]  # I列
        
        if method == 1:
            # 使用BCDEFG列
            if all(val in [0, 1] for val in [b_val, c_val, d_val, e_val, f_val, g_val, i_val]):
                X.append([b_val, c_val, d_val, e_val, f_val, g_val])
                y.append(i_val)
        elif method == 2:
            # 只使用BCD列
            if all(val in [0, 1] for val in [b_val, c_val, d_val, i_val]):
                X.append([b_val, c_val, d_val])
                y.append(i_val)
        elif method == 3:
            # 只使用EFG列
            if all(val in [0, 1] for val in [e_val, f_val, g_val, i_val]):
                X.append([e_val, f_val, g_val])
                y.append(i_val)
    
    X = np.array(X)
    y = np.array(y)
    
    print(f"数据加载完成，共{len(X)}个样本")
    print(f"特征数据形状: {X.shape}")
    print(f"标签数据形状: {y.shape}")
    

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)
    model = LogisticRegression()
    model.fit(X_train, y_train)
    
    y_pred = model.predict(X_test)
    y_pred_proba = model.predict_proba(X_test)
    
    print("\n模型评估结果:")
    print(f"准确率: {accuracy_score(y_test, y_pred):.4f}")
    print(f"F1分数: {f1_score(y_test, y_pred):.4f}")
    print(f"混淆矩阵:")
    print(confusion_matrix(y_test, y_pred))
    
    # 计算AUC-ROC
    if len(np.unique(y)) > 1:
        auc_roc = roc_auc_score(y_test, y_pred_proba[:, 1])
        print(f"AUC-ROC: {auc_roc:.4f}")
    
    print("\n模型系数:")
    print(f"截距: {model.intercept_[0]:.4f}")
    if method == 1:
        print(f"B列系数: {model.coef_[0][0]:.4f}")
        print(f"C列系数: {model.coef_[0][1]:.4f}")
        print(f"D列系数: {model.coef_[0][2]:.4f}")
        print(f"E列系数: {model.coef_[0][3]:.4f}")
        print(f"F列系数: {model.coef_[0][4]:.4f}")
        print(f"G列系数: {model.coef_[0][5]:.4f}")
    elif method == 2:
        print(f"B列系数: {model.coef_[0][0]:.4f}")
        print(f"C列系数: {model.coef_[0][1]:.4f}")
        print(f"D列系数: {model.coef_[0][2]:.4f}")
    elif method == 3:
        print(f"E列系数: {model.coef_[0][0]:.4f}")
        print(f"F列系数: {model.coef_[0][1]:.4f}")
        print(f"G列系数: {model.coef_[0][2]:.4f}")
    
    # 可视化ROC曲线和评估指标
    if len(np.unique(y)) > 1:
        fpr, tpr, thresholds = roc_curve(y_test, y_pred_proba[:, 1])
        plt.figure(figsize=(12, 6))
        
        # 绘制ROC曲线
        plt.subplot(1, 2, 1)
        plt.plot(fpr, tpr, label=f'ROC曲线 (AUC = {auc_roc:.4f})')
        plt.plot([0, 1], [0, 1], 'k--', label='随机猜测')
        
        # 找到最优准确率的点
        best_accuracy = 0
        best_threshold = 0
        best_fpr = 0
        best_tpr = 0
        
        for i, threshold in enumerate(thresholds):
            y_pred_threshold = (y_pred_proba[:, 1] >= threshold).astype(int)
            # 计算准确率
            accuracy = accuracy_score(y_test, y_pred_threshold)
            if accuracy > best_accuracy:
                best_accuracy = accuracy
                best_threshold = threshold
                best_fpr = fpr[i]
                best_tpr = tpr[i]
        
        # 标记最优准确率的点
        plt.scatter(best_fpr, best_tpr, color='red', s=50, label=f'最优阈值点（准确率={best_accuracy:.4f}）')
        
        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel('假阳性率')
        plt.ylabel('真阳性率')
        plt.title('AUC-ROC曲线')
        plt.legend(loc="lower right")
        
        # 显示评估指标
        plt.subplot(1, 2, 2)
        metrics_text = f"准确率: {accuracy_score(y_test, y_pred):.4f}\n"
        metrics_text += f"F1分数: {f1_score(y_test, y_pred):.4f}\n"
        metrics_text += f"AUC-ROC: {auc_roc:.4f}\n\n"

        metrics_text += f"模型系数:\n"
        metrics_text += f"截距: {model.intercept_[0]:.4f}\n"
        if method == 1:
            metrics_text += f"B列系数: {model.coef_[0][0]:.4f}\n"
            metrics_text += f"C列系数: {model.coef_[0][1]:.4f}\n"
            metrics_text += f"D列系数: {model.coef_[0][2]:.4f}\n"
            metrics_text += f"E列系数: {model.coef_[0][3]:.4f}\n"
            metrics_text += f"F列系数: {model.coef_[0][4]:.4f}\n"
            metrics_text += f"G列系数: {model.coef_[0][5]:.4f}\n\n"
        elif method == 2:
            metrics_text += f"B列系数: {model.coef_[0][0]:.4f}\n"
            metrics_text += f"C列系数: {model.coef_[0][1]:.4f}\n"
            metrics_text += f"D列系数: {model.coef_[0][2]:.4f}\n\n"
        elif method == 3:
            metrics_text += f"E列系数: {model.coef_[0][0]:.4f}\n"
            metrics_text += f"F列系数: {model.coef_[0][1]:.4f}\n"
            metrics_text += f"G列系数: {model.coef_[0][2]:.4f}\n\n"
        
        plt.text(0.1, 0.9, metrics_text, fontsize=10, verticalalignment='top')
        plt.axis('off')
        plt.title('模型评估指标')
        
        plt.tight_layout()
        plt.show()
    else:
        # 如果只有一个类别，只显示评估指标
        print("\n模型评估指标:")
        print(f"准确率: {accuracy_score(y_test, y_pred):.4f}")
        print(f"F1分数: {f1_score(y_test, y_pred):.4f}")
        print(f"\n模型系数:")
        print(f"截距: {model.intercept_[0]:.4f}")
        if method == 1:
            print(f"B列系数: {model.coef_[0][0]:.4f}")
            print(f"C列系数: {model.coef_[0][1]:.4f}")
            print(f"D列系数: {model.coef_[0][2]:.4f}")
            print(f"E列系数: {model.coef_[0][3]:.4f}")
            print(f"F列系数: {model.coef_[0][4]:.4f}")
            print(f"G列系数: {model.coef_[0][5]:.4f}")
        elif method == 2:
            print(f"B列系数: {model.coef_[0][0]:.4f}")
            print(f"C列系数: {model.coef_[0][1]:.4f}")
            print(f"D列系数: {model.coef_[0][2]:.4f}")
        elif method == 3:
            print(f"E列系数: {model.coef_[0][0]:.4f}")
            print(f"F列系数: {model.coef_[0][1]:.4f}")
            print(f"G列系数: {model.coef_[0][2]:.4f}")