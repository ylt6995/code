import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.metrics import roc_curve, roc_auc_score

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 文件路径
method = 1
result_path = '.\\实验结果\\output_method{}_gpt.xlsx'.format(method)
result_path2 = '.\\实验结果\\output_method{}.xlsx'.format(method)

bad_bids = [5, 7, 20, 34, 46, 54, 65, 76, 88, 99]
for i in range(len(bad_bids)):
    bad_bids[i] -= 1  # 调整为0-based索引

# 检查文件是否存在
if not os.path.exists(result_path):
    print(f"文件 {result_path} 不存在")
    exit()
if not os.path.exists(result_path2):
    print(f"文件 {result_path2} 不存在")
    exit()

# 加载第一个数据集 (GPT)
workbook1 = load_workbook(result_path)
worksheet1 = workbook1.active

row_numbers1 = []
scores1 = []
labels1 = []  # 0表示正常，1表示异常

for row in range(2, worksheet1.max_row + 1):
    row_numbers1.append(row)
    score = worksheet1.cell(row=row, column=2).value
    if score is None:
        score = 0
    scores1.append(score)
    
    # 标记样本：如果索引在bad_bids中则为异常(1)，否则为正常(0)
    idx = row - 2  # 转换为0-based索引
    labels1.append(1 if idx in bad_bids else 0)

# 分离第一个数据集的正常点和异常点
normal_row_numbers1 = []
normal_scores1 = []
abnormal_row_numbers1 = []
abnormal_scores1 = []

for idx, (row_num, score) in enumerate(zip(row_numbers1, scores1)):
    if idx in bad_bids:
        abnormal_row_numbers1.append(row_num)
        abnormal_scores1.append(score)
    else:
        normal_row_numbers1.append(row_num)
        normal_scores1.append(score)

# 计算第一个数据集的平均值
normal_avg1 = np.mean(normal_scores1) if normal_scores1 else 0
abnormal_avg1 = np.mean(abnormal_scores1) if abnormal_scores1 else 0

# 加载第二个数据集
workbook2 = load_workbook(result_path2)
worksheet2 = workbook2.active

row_numbers2 = []
scores2 = []
labels2 = []  # 0表示正常，1表示异常

for row in range(2, worksheet2.max_row + 1):
    row_numbers2.append(row)
    score = worksheet2.cell(row=row, column=2).value
    if score is None:
        score = 0
    scores2.append(score)
    
    # 标记样本：如果索引在bad_bids中则为异常(1)，否则为正常(0)
    idx = row - 2  # 转换为0-based索引
    labels2.append(1 if idx in bad_bids else 0)

# 分离第二个数据集的正常点和异常点
normal_row_numbers2 = []
normal_scores2 = []
abnormal_row_numbers2 = []
abnormal_scores2 = []

for idx, (row_num, score) in enumerate(zip(row_numbers2, scores2)):
    if idx in bad_bids:
        abnormal_row_numbers2.append(row_num)
        abnormal_scores2.append(score)
    else:
        normal_row_numbers2.append(row_num)
        normal_scores2.append(score)

# 计算第二个数据集的平均值
normal_avg2 = np.mean(normal_scores2) if normal_scores2 else 0
abnormal_avg2 = np.mean(abnormal_scores2) if abnormal_scores2 else 0

# 创建双图布局
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

# 左侧：散点图

point_size = 50
linewidth = 1
# 第一个数据集 (GPT)
ax1.scatter(normal_row_numbers1, normal_scores1, color='blue', alpha=0.6, s=point_size, label='正常招标 (GPT)')
ax1.scatter(abnormal_row_numbers1, abnormal_scores1, color='red', alpha=0.8, s=point_size, label='异常招标 (GPT)')
ax1.axhline(y=normal_avg1, color='blue', linestyle='--', linewidth=linewidth, alpha=0.7, label=f'正常项目平均分 (GPT): {normal_avg1:.2f}')
ax1.axhline(y=abnormal_avg1, color='red', linestyle='--', linewidth=linewidth, alpha=0.7, label=f'异常项目平均分 (GPT): {abnormal_avg1:.2f}')

# 第二个数据集
ax1.scatter(normal_row_numbers2, normal_scores2, color='blue', alpha=0.6, s=point_size, label='正常招标 (GLM)', marker='X', edgecolors='black')
ax1.scatter(abnormal_row_numbers2, abnormal_scores2, color='red', alpha=0.8, s=point_size, label='异常招标 (GLM)', marker='X', edgecolors='black')
ax1.axhline(y=normal_avg2, color='blue', linestyle='-', linewidth=linewidth, alpha=0.7, label=f'正常项目平均分 (GLM): {normal_avg2:.2f}')
ax1.axhline(y=abnormal_avg2, color='red', linestyle='-', linewidth=linewidth, alpha=0.7, label=f'异常项目平均分 (GLM): {abnormal_avg2:.2f}')

ax1.set_title('围标串标置信度散点分布图', fontsize=14, fontweight='bold')
ax1.set_xlabel('行号', fontsize=12)
ax1.set_ylabel('分数', fontsize=12)
ax1.grid(True, alpha=0.3)
ax1.legend()

# 右侧：AUC-ROC曲线
if len(set(labels1)) > 1:  # 确保第一个数据集有正样本和负样本
    fpr1, tpr1, _ = roc_curve(labels1, scores1)
    auc_score1 = roc_auc_score(labels1, scores1)
    ax2.plot(fpr1, tpr1, color='purple', lw=2, label=f'GPT, AUC = {auc_score1:.2f}')

if len(set(labels2)) > 1:  # 确保第二个数据集有正样本和负样本
    fpr2, tpr2, _ = roc_curve(labels2, scores2)
    auc_score2 = roc_auc_score(labels2, scores2)
    ax2.plot(fpr2, tpr2, color='orange', lw=2, label=f'GLM, AUC = {auc_score2:.2f}')

ax2.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--', label='随机猜测')
ax2.set_xlim([0.0, 1.0])
ax2.set_ylim([0.0, 1.01])
ax2.set_xlabel('假阳性率', fontsize=12)
ax2.set_ylabel('真阳性率', fontsize=12)
ax2.set_title('AUC-ROC曲线', fontsize=14, fontweight='bold')
ax2.grid(True, alpha=0.3)
ax2.legend(loc="lower right")

# 调整布局
plt.tight_layout()

# 显示图表
plt.show()

# 关闭工作簿
workbook1.close()
workbook2.close()