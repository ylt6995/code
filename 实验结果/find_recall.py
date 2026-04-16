import openpyxl

# 你的结果文件路径
result_path = r"实验结果\output_method3_model-zhipu_masked.xlsx"

# 阈值：模型分数大于等于多少，判定为围标
THRESHOLD = 80

# 统计变量
TP = 0  # 预测围标 + 真实围标 ✅
FP = 0  # 预测围标 + 真实非围标 ❌
FN = 0  # 预测非围标 + 真实围标 ❌
TN = 0  # 预测非围标 + 真实非围标 ✅

# 读取结果
wb = openpyxl.load_workbook(result_path)
ws = wb.active

# 从第2行开始遍历（第1行是表头）
for row in range(2, ws.max_row + 1):
    # B列：模型分数（collusionSuspicionScore）
    score = float(ws.cell(row=row, column=2).value or 0)
    # I列：真实标签（1=围标，0=非围标）
    true_label = int(ws.cell(row=row, column=9).value or 0)
    
    # 模型预测结果
    pred_label = 1 if score >= THRESHOLD else 0

    # 计算 TP/FP/FN/TN
    if pred_label == 1 and true_label == 1:
        TP += 1
    elif pred_label == 1 and true_label == 0:
        FP += 1
    elif pred_label == 0 and true_label == 1:
        FN += 1
    elif pred_label == 0 and true_label == 0:
        TN += 1

# ==================== 输出指标 ====================
print(f"score≥{THRESHOLD})")
print(f"TP ：{TP}")
print(f"FP：{FP}")
print(f"FN ：{FN}")
print(f"TN ：{TN}")
print("\n===== core =====")
precision = TP / (TP + FP) if (TP + FP) > 0 else 0
recall = TP / (TP + FN) if (TP + FN) > 0 else 0
f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

print(f"Precision：{precision:.2f}")
print(f"Recall：{recall:.2f}")
print(f"F1：{f1:.2f}")
print(f"TP+FN：{TP + FN}")
print(f"TP：{TP}")