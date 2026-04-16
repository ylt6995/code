import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.metrics import roc_curve, roc_auc_score

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 文件路径
result_paths = ['.\实验结果\output_method1.xlsx', '.\实验结果\output_method2.xlsx', '.\实验结果\output_method3.xlsx']
method_names = ['无提示', '少样本提示', '思维链']
colors = ['orange', 'red', 'blue']

bad_bids = [5, 7, 20, 34, 46, 54, 65, 76, 88, 99]
for i in range(len(bad_bids)):
    bad_bids[i] -= 1  # 调整为0-based索引

# 检查所有文件是否存在
for path in result_paths:
    if not os.path.exists(path):
        print(f"文件 {path} 不存在")
        exit()

# 加载并处理所有数据集
datasets = []
for path in result_paths:
    workbook = load_workbook(path)
    worksheet = workbook.active
    
    scores = []
    labels = []  # 0表示正常，1表示异常
    
    for row in range(2, worksheet.max_row + 1):
        score = worksheet.cell(row=row, column=2).value
        if score is None:
            score = 0
        scores.append(score)
        
        # 标记样本：如果索引在bad_bids中则为异常(1)，否则为正常(0)
        idx = row - 2  # 转换为0-based索引
        labels.append(1 if idx in bad_bids else 0)
    
    datasets.append((scores, labels, workbook))

# 创建ROC图
plt.figure(figsize=(10, 8))

# 绘制每组数据的ROC曲线
for i, (scores, labels, workbook) in enumerate(datasets):
    if len(set(labels)) > 1:  # 确保数据集有正样本和负样本
        fpr, tpr, _ = roc_curve(labels, scores)
        auc_score = roc_auc_score(labels, scores)
        plt.plot(fpr, tpr, color=colors[i], lw=2, label=f'{method_names[i]}, AUC = {auc_score:.2f}')
    workbook.close()

# 绘制随机猜测线
plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--', label='随机猜测')

# 设置图表属性
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.01])
plt.xlabel('假阳性率', fontsize=12)
plt.ylabel('真阳性率', fontsize=12)
plt.title('三中方法的ROC曲线对比', fontsize=14, fontweight='bold')
plt.grid(True, alpha=0.3)
plt.legend(loc="lower right")

# 调整布局
plt.tight_layout()

# 显示图表
plt.show()