import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from sklearn.metrics import roc_curve, roc_auc_score, accuracy_score

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 文件路径
method = 3
company = "zhipu"   # "zhipu" 或 "chat"
result_path = '.\\实验结果\\output_method{}_model-{}_masked.xlsx'.format(method, company)
label_path = '.\\数据\\testdata_masked.xlsx'

bad_bids = []
w1 = load_workbook(label_path)
ws1 = w1.active
bad_table = []
for cell in ws1['B']:
    if cell.value != None:
        bad_table.append(int(cell.value))
for i in range(len(bad_table)):
    if bad_table[i] == 1:
        bad_bids.append(i+1)  # 转换为0-based索引
print(f"异常招标索引: {bad_bids}, 共{len(bad_bids)}个")

# 检查文件是否存在
if not os.path.exists(result_path):
    print(f"文件 {result_path} 不存在")
    exit()

# 加载数据
workbook = load_workbook(result_path)
worksheet = workbook.active

row_numbers = []
scores = []
labels = []  # 0表示正常，1表示异常

for row in range(2, worksheet.max_row + 1):
    row_numbers.append(row)
    score = worksheet.cell(row=row, column=2).value
    if score is None:
        score = 0
    scores.append(score)
    
    # 标记样本：如果索引在bad_bids中则为异常(1)，否则为正常(0)
    idx = row - 2  # 转换为0-based索引
    labels.append(1 if idx in bad_bids else 0)

# 分离正常点和异常点
normal_row_numbers = []
normal_scores = []
abnormal_row_numbers = []
abnormal_scores = []

for idx, (row_num, score) in enumerate(zip(row_numbers, scores)):
    if idx in bad_bids:
        abnormal_row_numbers.append(row_num)
        abnormal_scores.append(score)
    else:
        normal_row_numbers.append(row_num)
        normal_scores.append(score)

# 计算平均值
normal_avg = np.mean(normal_scores) if normal_scores else 0
abnormal_avg = np.mean(abnormal_scores) if abnormal_scores else 0

# 创建双图布局
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

# 左侧：散点图
ax1.scatter(normal_row_numbers, normal_scores, color='blue', alpha=0.6, s=50, label='正常招标')
ax1.scatter(abnormal_row_numbers, abnormal_scores, color='red', alpha=0.8, s=70, label='异常招标')
ax1.axhline(y=normal_avg, color='blue', linestyle='--', linewidth=2, alpha=0.7, label=f'正常项目平均分: {normal_avg:.2f}')
ax1.axhline(y=abnormal_avg, color='red', linestyle='--', linewidth=2, alpha=0.7, label=f'异常项目平均分: {abnormal_avg:.2f}')
ax1.set_title('招标分数散点图', fontsize=14, fontweight='bold')
ax1.set_xlabel('行号', fontsize=12)
ax1.set_ylabel('分数', fontsize=12)
ax1.grid(True, alpha=0.3)
ax1.legend()

# 右侧：AUC-ROC曲线
if len(set(labels)) > 1:  # 确保有正样本和负样本
    fpr, tpr, thresholds = roc_curve(labels, scores)
    auc_score = roc_auc_score(labels, scores)
    
    # 计算最优准确率
    best_accuracy = 0
    best_threshold = 0
    best_fpr = 0
    best_tpr = 0
    
    for i, threshold in enumerate(thresholds):
        # 根据阈值预测标签
        y_pred = [1 if score >= threshold else 0 for score in scores]
        # 计算准确率
        accuracy = accuracy_score(labels, y_pred)
        if accuracy > best_accuracy:
            best_accuracy = accuracy
            best_threshold = threshold
            best_fpr = fpr[i]
            best_tpr = tpr[i]
    
    ax2.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC曲线 (AUC = {auc_score:.2f})')
    ax2.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--', label='随机猜测')
    # 标记最优准确率点
    ax2.scatter(best_fpr, best_tpr, color='green', s=100, marker='o', label=f'最优阈值点（准确率={best_accuracy:.2f}）')
    ax2.set_xlim([0.0, 1.0])
    ax2.set_ylim([0.0, 1.01])
    ax2.set_xlabel('假阳性率', fontsize=12)
    ax2.set_ylabel('真阳性率', fontsize=12)
    ax2.set_title('AUC-ROC曲线', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3)
    ax2.legend(loc="lower right")
else:
    ax2.text(0.5, 0.5, '无法计算ROC曲线\n（需要同时包含正样本和负样本）', 
             ha='center', va='center', fontsize=12)
    ax2.set_title('AUC-ROC曲线', fontsize=14, fontweight='bold')

# 调整布局
plt.tight_layout()

# 显示图表
plt.show()

# 关闭工作簿
workbook.close()