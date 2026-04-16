import openpyxl
import json
import openai
from time import sleep, time
import os
from datetime import datetime
import re

company = os.getenv("LLM_PROVIDER", "zhipu")
model_list = {"zhipu": os.getenv("ZHIPU_MODEL", "glm-4.7-flash"), "chat": os.getenv("GPT_MODEL", "gpt-4o-mini")}

zhipu_api_key = os.getenv("ZHIPU_API_KEY", "")
zhipu_base_url = os.getenv("ZHIPU_BASE_URL", "https://open.bigmodel.cn/api/paas/v4")
chat_api_key = os.getenv("GPT_API_KEY", "") or os.getenv("OPENAI_API_KEY", "")
chat_base_url = os.getenv("GPT_BASE_URL", "") or os.getenv("OPENAI_BASE_URL", "https://api.chatanywhere.tech/v1")

ak = eval(f"{company}_api_key")
bu = eval(f"{company}_base_url")
model_name = model_list[company]

method = 3 # 1, 2, 3分别对应三种提示词设计
delay = 20  # 每次请求的延迟时间，单位秒

# Excel文件路径
excel_path = '.\\data\\testdata_masked.xlsx'
result_path = '.\\实验结果\\output_method{}_model-{}_masked.xlsx'.format(method, company)

# 加载Excel文件
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 读取所有单元格内容
rows = list(ws['A'])


bad_bids = []
bad_table = []
for cell in ws['B']:
    if cell.value != None:
        bad_table.append(int(cell.value))
for i in range(len(bad_table)):
    if bad_table[i] == 1:
        bad_bids.append(i+1)  # 转换为0-based索引
print(f"异常招标索引: {bad_bids}, 共{len(bad_bids)}个")

# 分类存储
xm_list, zb_list, tb_list = [], [], []
connecter = ["projguid", "bid_ann_guid"]

for cell in rows:
    if not cell.value or cell.value[0] != "{":
        continue
    try:
        data_dict = json.loads(cell.value)
        if connecter[0] in data_dict and connecter[1] not in data_dict:
            xm_list.append(data_dict)
        elif connecter[1] in data_dict and connecter[0] not in data_dict:
            tb_list.append(data_dict)
        elif connecter[0] in data_dict and connecter[1] in data_dict:
            zb_list.append(data_dict)
    except Exception as e:
        continue

def check_collusion(prompt):
    sleep(delay)  # 避免请求过于频繁
    client = openai.OpenAI(api_key = ak, base_url = bu)
    while True:
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[{"role": "system", "content": "You are an expert in procurement auditing and bid rigging detection."},
                          {"role": "user", "content": prompt}],
                timeout=300.0  # 5分钟超时
            )
            return response.choices[0].message.content
        except Exception as e:
            # 检查是否是超时异常
            if "timeout" in str(e).lower() or "timed out" in str(e).lower():
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 请求超时，正在重试...")
                sleep(delay)  # 再次延迟，避免频繁请求
                continue
            else:
                # 其他异常，直接抛出
                raise

# 检查结果文件是否存在
processed_bid_ann_guids = set()

if os.path.exists(result_path):
    # 读取已处理的bid_ann_guid
    result_wb = openpyxl.load_workbook(result_path)
    result_ws = result_wb.active
    for row in result_ws.iter_rows(min_row=2, max_row=result_ws.max_row, min_col=1, max_col=1):
        bid_ann_guid = row[0].value
        if bid_ann_guid:
            processed_bid_ann_guids.add(bid_ann_guid)
else:
    # 创建新的结果文件
    result_wb = openpyxl.Workbook()
    result_ws = result_wb.active
    # 设置表头
    headers = ["bid_ann_guid", "collusionSuspicionScore", "riskLevel", "keyEvidence"]
    for col, header in enumerate(headers, 1):
        result_ws.cell(row=1, column=col, value=header)

# 计算总项目数和已完成项目数
total_projects = len(zb_list)
completed_projects = len(processed_bid_ann_guids)

# 打印提示词，仅一次
prompt_show_flag = True

# 遍历每个招标，提取对应的五个投标
for zb in zb_list:
    current_start_time = time()
    bid_ann_guid = zb['bid_ann_guid']
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 检测结果:\n")
    # 跳过已处理的项目
    if bid_ann_guid in processed_bid_ann_guids:
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 已处理，跳过 | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)")
        continue
    
    # tb_group = [tb for tb in tb_list if tb['bid_ann_guid'] == bid_ann_guid]
    tb_group = []
    for tb in tb_list:
        if tb['bid_ann_guid'] == bid_ann_guid:
            # 选择特定的字段
            selected_tb = {}
            selected_tb['x_providername'] = tb.get('x_providername', '-1')
            selected_tb['x_price'] = tb.get('x_price', -1)
            selected_tb['versionnumber'] = tb.get('versionnumber', -1)
            selected_tb['x_winsTheBid'] = tb.get('x_isqualified', -1)
            tb_group.append(selected_tb)
    if len(tb_group) != 5:
        assert(False)

    # 构造提示词
    prompt = "\n"
    for idx, tb in enumerate(tb_group, 1):
        prompt += f"Bidding information{idx}: {json.dumps(tb, ensure_ascii=False)}\n"

    # 方法1
    prompt1 = f"""Analyze this list of bidders for collusion (bid-rigging) risk and output a JSON object with the following keys: "collusionSuspicionScore" (0-100), "riskLevel", "keyEvidence".\
    Bidder Data:'''{prompt}'''\n"""
    # 方法2
    prompt2 = f"""Output Format Requirements:
    Always output a valid JSON object with exactly these keys:
    - collusionSuspicionScore (0-100)
    - riskLevel ("Low", "Medium", "High", "Critical")
    - keyEvidence (array of concise strings)
    - mostSuspiciousBidders (array of x_providername)

Example - Data:
```
[
  {{"x_providername": "Delta Designs", "x_price": 50000, "x_biddercontact": "16666666666"}},
  {{"x_providername": "Epsilon Builders", "x_price": 98000, "x_biddercontact": "17777777777"}},
  {{"x_providername": "Zeta Constructors", "x_price": 99000, "x_biddercontact": "18888888888"}}
]
```

Example – Output:
```
{{
  "collusionSuspicionScore": 80,
  "riskLevel": "High",
  "keyEvidence": ["One bid is significantly lower than the others, creating a price anchor.", "The two highest bids are unusually close in value."],
  "mostSuspiciousBidders": ["Epsilon Builders", "Zeta Constructors"]
}}
```

Now, analyze the following new bidder data:'''{prompt}'''\n
"""
    # 方法3
    prompt3 = f"""You need to analyze the provided bidder data for a single tender project and perform a comprehensive collusion risk assessment.
Core Task:
    Identify red flags and calculate a Collusion Suspicion Score (0-100) based on the evidence found. Focus on hard evidence and behavioral patterns.

Analysis Process:
    1. Is there an abnormally low bid alongside a cluster of closely priced high bids? Do the bid prices show a mathematical relationship or consistent gaps?
    2. Are there patterns directly indicating collusion and bid-rigging cartels? or are they simply manifestations of other unrelated behaviors?
    3. Based on your previous analysis, assign a Collusion Suspicion Score (0-100) and categorize the risk level into "Low", "Medium", "High", or "Critical".

Output Format Requirements:
    Provide your analysis in the following structured JSON format:
```
{{
  "riskLevel": ["Low" | "Medium" | "High" | "Critical"],
  "keyEvidence": [
    {{"type": "[Pricing Pattern]", "description": "Clear description of the finding"}}
  ],
  "mostSuspiciousGroup": {{
    "bidders": ["List of provider names forming the most likely collusion group"],
    "reasoning": "Explanation for why this group is particularly suspect"
  }},
  "recommendedInvestigationSteps": ["A prioritized list of 2-3 concrete next steps for audit"],
  "collusionSuspicionScore": [A number between 0-100]
}}
```

Data for Analysis: ```{prompt}```\n
"""
    
    prompt = eval(f"prompt{method}")

    # 打印提示词，仅一次
    if prompt_show_flag:
        print(f"提示词示例 (方法{method}):\n{prompt}\n")
        prompt_show_flag = False


    # 检测是否存在围标串标
    result = check_collusion(prompt)
    print(f"{result}")
    
    # 解析结果
    try:
        # 提取JSON部分
        while True:
            try:
                json_match = re.search(r'\{.*\}', result, re.DOTALL)
                if json_match:
                    result_json = json.loads(json_match.group())
                else:
                    # 如果没有找到JSON，尝试直接解析
                    result_json = json.loads(result)
            except:
                pass
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析JSON失败，等待120秒后重试")
                sleep(120)
                result = check_collusion(prompt)
            else:
                break
        
        # 提取数据
        collusion_score = result_json.get("collusionSuspicionScore", "")
        risk_level = result_json.get("riskLevel", "")
        key_evidence = result_json.get("keyEvidence", "")

        # 计算处理耗时（从上一行写入到这一行写入之间的时间差）
        current_end_time = time()
        processing_time = current_end_time - current_start_time - delay  # 减去请求延迟时间
        
        # 写入结果到Excel
        next_row = result_ws.max_row + 1
        result_ws.cell(row=next_row, column=1, value=bid_ann_guid)
        result_ws.cell(row=next_row, column=2, value=collusion_score)
        result_ws.cell(row=next_row, column=3, value=str(risk_level))
        result_ws.cell(row=next_row, column=4, value=str(key_evidence))
        result_ws.cell(row=next_row, column=5, value=processing_time)

        # 写入mostSuspiciousBidders
        try:
            most_suspicious_bidders = result_json.get("mostSuspiciousGroup", "")
            result_ws.cell(row=next_row, column=6, value=str(most_suspicious_bidders))
        except:
            pass

        try:
            recommended_steps = result_json.get("recommendedInvestigationSteps", "")
            result_ws.cell(row=next_row, column=7, value=str(recommended_steps))
        except:
            pass

        result_ws.cell(row=next_row, column=8, value=str(prompt))
        result_ws.cell(row=next_row, column=9, value=bad_table.pop(0))
        
        # 保存结果文件
        result_wb.save(result_path)
        completed_projects += 1
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 结果已保存到{result_path} | 处理耗时: {processing_time:.2f}秒(不计{delay}秒延迟) | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)\n")
        
    except Exception as e:
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 招标编号: {bid_ann_guid} 解析结果失败: {e} | 进度: {completed_projects}/{total_projects} ({completed_projects/total_projects*100:.1f}%)")
        #print(json_match.group())
        assert(False)

# 关闭结果文件
result_wb.close()
